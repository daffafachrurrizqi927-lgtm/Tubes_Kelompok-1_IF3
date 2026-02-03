from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import pandas as pd
import re 
from datetime import datetime
from openpyxl import load_workbook 
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# --- DAFTAR 48 KOTA VALID (DATA ASLI) ---
locations_list = [
    ("Banda Aceh", "indonesia/banda-aceh", 5.5483, 95.3238),
    ("Medan", "indonesia/medan", 3.5952, 98.6722),
    ("Pematangsiantar", "indonesia/pematangsiantar", 2.9667, 99.0667),
    ("Padang", "indonesia/padang", -0.9471, 100.4172),
    ("Bukittinggi", "indonesia/bukittinggi", -0.3000, 100.3833),
    ("Pekanbaru", "indonesia/pekanbaru", 0.5071, 101.4478),
    ("Jambi", "indonesia/jambi", -1.6101, 103.6197),
    ("Palembang", "indonesia/palembang", -2.9761, 104.7754),
    ("Bengkulu", "indonesia/bengkulu", -3.8004, 102.2655),
    ("Pangkal Pinang", "indonesia/pangkal-pinang", -2.1333, 106.1167),
    ("Tanjung Pinang", "indonesia/tanjung-pinang", 0.9167, 104.4500),
    ("Batam", "indonesia/batam", 1.0456, 104.0305),
    ("Jakarta", "indonesia/jakarta", -6.2088, 106.8456),
    ("Serang", "indonesia/serang", -6.1200, 106.1503),
    ("Bekasi", "indonesia/bekasi", -6.2383, 106.9756),
    ("Bogor", "indonesia/bogor", -6.5971, 106.8060),
    ("Bandung", "indonesia/bandung", -6.9175, 107.6191),
    ("Tasikmalaya", "indonesia/tasikmalaya", -7.3274, 108.2207),
    ("Cirebon", "indonesia/cirebon", -6.7320, 108.5523),
    ("Semarang", "indonesia/semarang", -6.9667, 110.4167),
    ("Tegal", "indonesia/tegal", -6.8694, 109.1402),
    ("Pekalongan", "indonesia/pekalongan", -6.8898, 109.6746),
    ("Kudus", "indonesia/kudus", -6.8048, 110.8405),
    ("Surakarta", "indonesia/surakarta", -7.5755, 110.8243),
    ("Yogyakarta", "indonesia/yogyakarta", -7.7956, 110.3695),
    ("Surabaya", "indonesia/surabaya", -7.2575, 112.7521),
    ("Malang", "indonesia/malang", -7.9666, 112.6326),
    ("Madiun", "indonesia/madiun", -7.6298, 111.5177),
    ("Kediri", "indonesia/kediri", -7.8485, 112.0183),
    ("Pontianak", "indonesia/pontianak", -0.0263, 109.3425),
    ("Palangkaraya", "indonesia/palangkaraya", -2.2083, 113.9167),
    ("Banjarmasin", "indonesia/banjarmasin", -3.3167, 114.5928),
    ("Balikpapan", "indonesia/balikpapan", -1.2675, 116.8289),
    ("Samarinda", "indonesia/samarinda", -0.5022, 117.1536),
    ("Makassar", "indonesia/makassar", -5.1477, 119.4327),
    ("Palu", "indonesia/palu", -0.8917, 119.8707),
    ("Kendari", "indonesia/kendari", -3.9972, 122.5120),
    ("Gorontalo", "indonesia/gorontalo", 0.5412, 123.0595),
    ("Manado", "indonesia/manado", 1.4748, 124.8421),
    ("Denpasar", "indonesia/denpasar", -8.6705, 115.2126),
    ("Singaraja", "indonesia/singaraja", -8.1120, 115.0882),
    ("Mataram", "indonesia/mataram", -8.5833, 116.1167),
    ("Kupang", "indonesia/kupang", -10.1772, 123.6070),
    ("Ambon", "indonesia/ambon", -3.6954, 128.1814),
    ("Ternate", "indonesia/ternate", 0.7833, 127.3667),
    ("Sofifi", "indonesia/sofifi", 0.7200, 127.5700),
    ("Manokwari", "indonesia/manokwari", -0.8614, 134.0620),
    ("Jayapura", "indonesia/jayapura", -2.5337, 140.7181)
]

locations = {name: {"slug": slug, "lat": lat, "lon": lon} for name, slug, lat, lon in locations_list}
all_data = []

print(f"🤖 MENYIAPKAN SELENIUM (KHUSUS PERSENTASE %)...")

chrome_options = Options()
# Set bahasa Inggris agar format angka standar
chrome_options.add_argument("--lang=en-US") 
# Paksa resolusi layar besar agar kolom tidak tersembunyi
chrome_options.add_argument("--window-size=1920,1080")

prefs = {"profile.managed_default_content_settings.images": 2}
chrome_options.add_experimental_option("prefs", prefs)
chrome_options.page_load_strategy = 'eager'
chrome_options.add_argument("--disable-blink-features=AutomationControlled") 
chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36")

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
driver.set_page_load_timeout(30)

print("🚀 MEMULAI SCRAPING...")

try:
    wait = WebDriverWait(driver, 10) 

    for i, (city_name, info) in enumerate(locations.items(), 1):
        print(f"[{i}/{len(locations)}] 🌍 {city_name}...", end=" ")
        
        url = f"https://www.timeanddate.com/weather/{info['slug']}/hourly"
        
        try:
            driver.get(url)
            try:
                wait.until(EC.presence_of_element_located((By.ID, "wt-hbh")))
            except:
                print("❌ Lewati (Timeout)")
                continue

            soup = BeautifulSoup(driver.page_source, 'html.parser')
            target_table = soup.find('table', id='wt-hbh')
            
            if target_table:
                data_count = 0
                rows = target_table.find('tbody').find_all('tr')
                current_date_str = datetime.now().strftime("%Y-%m-%d")

                for row in rows:
                    # 1. HEADER TANGGAL
                    header_col = row.find('th', colspan=True)
                    if header_col:
                        try:
                            date_text = header_col.get_text(strip=True)
                            if "," in date_text:
                                date_part = date_text.split(",")[1].strip()
                                dt_temp = datetime.strptime(date_part + f" {datetime.now().year}", "%d %B %Y")
                                current_date_str = dt_temp.strftime("%Y-%m-%d")
                        except: pass
                        continue

                    # 2. DATA JAM
                    cols = row.find_all('td')
                    if len(cols) > 8: 
                        th = row.find('th')
                        if not th: continue
                        clean_time = th.get_text(strip=True)[:5] # Ambil HH:MM
                        
                        deskripsi = cols[0].get_text(strip=True) 
                        if not deskripsi: 
                            img = cols[0].find('img')
                            if img: deskripsi = img.get('title', 'N/A')
                        
                        # --- FOKUS: AMBIL PERSENTASE (%) ---
                        peluang_hujan = "0%" # Default
                        
                        for col in cols:
                            txt = col.get_text(strip=True)
                            
                            # Cari angka yang diikuti tanda %
                            # Contoh teks di web: "24%" atau "24% 0.5mm"
                            if "%" in txt:
                                match_pct = re.search(r"(\d+)%", txt)
                                if match_pct:
                                    peluang_hujan = f"{match_pct.group(1)}%"
                                    break # Ketemu, langsung keluar loop

                        all_data.append({
                            "Kota": city_name, 
                            "Latitude": info['lat'], "Longitude": info['lon'],
                            "Tipe_Data": "Per Jam (48 Jam)",
                            "Tanggal": current_date_str,
                            "Jam": clean_time,
                            "Deskripsi": deskripsi, 
                            "Peluang_Hujan": peluang_hujan # HANYA PERSEN
                        })
                        data_count += 1
                
                print(f"✅ OK ({data_count} jam data)")
            else:
                print("⚠️ Gagal Baca Tabel")
        
        except Exception as e:
            print(f"❌ Error: {e}")
            continue

finally:
    try: driver.quit()
    except: pass

# --- SIMPAN EXCEL ---
print("\n" + "="*40)
df = pd.DataFrame(all_data)
filename = "Data_Cuaca.xlsx"

if not df.empty:
    df.to_excel(filename, index=False, engine='openpyxl')
    
    wb = load_workbook(filename)
    ws = wb.active
    header_fill = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for column_cells in ws.columns:
        length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

    wb.save(filename)
    print(f"✅ FILE SELESAI: {filename}")
    print(f"🎉 Total Data Valid (Persentase Only): {len(df)} Baris")
else:
    print("❌ Gagal mendapatkan data.")